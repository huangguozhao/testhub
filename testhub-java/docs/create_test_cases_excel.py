#!/usr/bin/env python3
"""
生成TestHub Java版本系统测试用例Excel
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 创建工作簿
wb = openpyxl.Workbook()

# 定义样式
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

sub_header_fill = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

wrap_alignment = Alignment(vertical="top", wrap_text=True)
center_alignment = Alignment(horizontal="center", vertical="center")

# P0优先级填充色
p0_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # 黄色
p1_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 绿色
p2_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # 蓝色

# 定义测试用例数据
test_cases = {
    "一、认证模块": [
        ("AUTH-001", "正常注册", "无", "输入用户名、邮箱、密码，点击注册", "注册成功，跳转登录页，收到Token", "P0"),
        ("AUTH-002", "用户名已存在", "用户名已被注册", "使用已存在的用户名注册", "提示"用户名已存在"，注册失败", "P0"),
        ("AUTH-003", "邮箱已注册", "邮箱已被使用", "使用已存在的邮箱注册", "提示"邮箱已被注册"，注册失败", "P0"),
        ("AUTH-004", "用户名为空", "无", "用户名留空，填写其他必填项", "提示"用户名不能为空"", "P0"),
        ("AUTH-005", "邮箱格式错误", "无", "输入格式错误的邮箱", "提示"邮箱格式不正确"", "P1"),
        ("AUTH-006", "密码过短", "无", "密码少于6位", "提示"密码至少6位"", "P1"),
        ("AUTH-007", "注册后自动登录", "无", "完成注册", "返回Token和用户信息", "P2"),
        ("AUTH-101", "正常登录", "用户已注册", "输入正确的用户名密码", "登录成功，返回Token，跳转首页", "P0"),
        ("AUTH-102", "用户名错误", "无", "输入不存在的用户名", "提示"用户名或密码错误"", "P0"),
        ("AUTH-103", "密码错误", "用户已注册", "输入正确的用户名，错误的密码", "提示"用户名或密码错误"", "P0"),
        ("AUTH-104", "用户已禁用", "用户状态为disabled", "使用该用户登录", "提示"用户已被禁用"", "P0"),
        ("AUTH-105", "用户名为空", "无", "用户名留空", "提示"用户名不能为空"", "P1"),
        ("AUTH-106", "密码为空", "无", "密码留空", "提示"密码不能为空"", "P1"),
        ("AUTH-107", "Token过期", "AccessToken已过期", "使用过期Token访问API", "返回401，自动刷新或跳转登录", "P1"),
        ("AUTH-108", "RefreshToken过期", "RefreshToken已过期", "Token过期后刷新", "提示"刷新令牌已过期"，跳转登录", "P0"),
        ("AUTH-201", "正常登出", "已登录", "点击登出按钮", "清除本地Token，跳转登录页", "P0"),
        ("AUTH-202", "登出后访问受限资源", "已登出", "使用登出后的Token访问API", "返回401，跳转登录页", "P0"),
        ("AUTH-203", "重复登出", "已登出", "再次点击登出", "无异常，不重复请求", "P1"),
        ("AUTH-301", "正常刷新", "已登录，Token即将过期", "等待Token即将过期", "自动刷新Token", "P0"),
        ("AUTH-302", "使用RefreshToken刷新", "AccessToken过期", "调用刷新接口", "返回新的AccessToken", "P0"),
        ("AUTH-303", "RefreshToken无效", "无", "使用伪造的RefreshToken", "提示"刷新令牌无效"", "P0"),
        ("AUTH-304", "RefreshToken已使用", "已使用过的RefreshToken", "再次使用该Token", "提示"刷新令牌已失效"", "P0"),
    ],
    "二、项目管理": [
        ("PROJ-001", "查看项目列表", "已登录", "进入项目管理页面", "显示当前用户参与的所有项目", "P0"),
        ("PROJ-002", "项目分页", "项目数量超过分页大小", "查看分页控件", "分页正常，支持切换页码", "P1"),
        ("PROJ-003", "项目搜索", "存在多个项目", "输入项目名称搜索", "显示匹配的项目", "P1"),
        ("PROJ-004", "无项目时显示", "用户未创建任何项目", "进入项目管理页", "显示空状态提示", "P2"),
        ("PROJ-101", "正常创建项目", "无", "填写项目名称、描述，创建", "项目创建成功，出现在列表中", "P0"),
        ("PROJ-102", "项目名称为空", "无", "项目名称留空", "提示"项目名称不能为空"", "P0"),
        ("PROJ-103", "项目名称重复", "已存在同名项目", "创建同名项目", "提示"项目名称已存在"", "P1"),
        ("PROJ-104", "项目名称超长", "无", "输入超长项目名称(>100字符)", "提示名称过长或截断", "P1"),
        ("PROJ-105", "描述超长", "无", "输入超长描述(>500字符)", "提示描述过长或截断", "P2"),
        ("PROJ-201", "正常编辑", "项目存在", "修改项目信息并保存", "保存成功，列表显示更新", "P0"),
        ("PROJ-202", "编辑权限", "非项目成员", "尝试编辑他人项目", "提示无权限", "P0"),
        ("PROJ-203", "删除项目", "项目存在", "点击删除项目", "弹出确认框，确认后删除", "P0"),
        ("PROJ-204", "删除后恢复", "项目已删除", "尝试访问被删除项目", "返回404", "P1"),
        ("PROJ-301", "添加成员", "项目存在", "添加新成员到项目", "成员添加成功", "P0"),
        ("PROJ-302", "移除成员", "成员已存在", "移除项目成员", "成员移除成功", "P1"),
        ("PROJ-303", "添加自己", "无", "尝试将自己添加为成员", "提示不能添加自己", "P2"),
        ("PROJ-304", "重复添加", "成员已存在", "再次添加该成员", "提示成员已存在", "P2"),
    ],
    "三、API Testing模块": [
        ("ENV-001", "创建环境", "项目存在", "创建测试环境，配置变量", "环境创建成功", "P0"),
        ("ENV-002", "环境变量使用", "环境已创建", "在环境变量中使用{{variable}}", "变量被正确替换", "P0"),
        ("ENV-003", "全局参数", "环境已创建", "配置全局参数", "所有请求可使用全局参数", "P1"),
        ("ENV-004", "环境切换", "多个环境", "在不同环境间切换", "URL和变量正确切换", "P0"),
        ("ENV-005", "变量覆盖", "环境变量有多个来源", "测试优先级", "高优先级变量覆盖低优先级", "P1"),
        ("ENV-006", "环境列表", "项目存在", "查看环境列表", "显示所有环境", "P0"),
        ("ENV-007", "编辑环境", "环境存在", "修改环境配置", "修改成功", "P0"),
        ("ENV-008", "删除环境", "环境存在", "删除环境", "确认后删除成功", "P0"),
        ("ENV-009", "删除被引用的环境", "环境被接口引用", "删除该环境", "提示环境被引用，无法删除", "P1"),
        ("API-001", "创建GET接口", "项目/集合存在", "创建GET请求接口", "接口创建成功", "P0"),
        ("API-002", "创建POST接口", "项目/集合存在", "创建POST请求，配置JSON Body", "接口创建成功", "P0"),
        ("API-003", "创建带Header接口", "项目/集合存在", "配置自定义Header", "Header正确发送", "P0"),
        ("API-004", "创建WebSocket接口", "项目/集合存在", "创建WebSocket类型接口", "接口创建成功", "P1"),
        ("API-005", "接口参数化", "环境变量已配置", "使用{{variable}}作为URL参数", "变量被正确替换", "P0"),
        ("API-006", "执行接口", "接口已创建", "点击发送按钮执行接口", "显示请求结果", "P0"),
        ("API-007", "响应时间显示", "接口已执行", "查看执行结果", "显示响应时间", "P1"),
        ("API-008", "状态码显示", "接口已执行", "查看执行结果", "显示HTTP状态码", "P0"),
        ("API-009", "响应体显示", "接口已执行", "查看响应内容", "正确显示JSON/文本", "P0"),
        ("API-010", "响应头显示", "接口已执行", "查看响应头", "显示所有响应头", "P2"),
        ("API-011", "执行失败-网络错误", "服务器不可达", "执行一个指向无效URL的请求", "显示网络错误信息", "P0"),
        ("API-012", "执行失败-超时", "服务器响应慢", "设置短超时，执行慢请求", "显示超时错误", "P1"),
        ("API-013", "执行失败-SSL错误", "证书问题", "执行HTTPS请求到证书问题站点", "显示SSL错误", "P1"),
        ("API-014", "编辑接口", "接口存在", "修改接口配置", "修改成功", "P0"),
        ("API-015", "复制接口", "接口存在", "复制接口", "生成新的接口副本", "P1"),
        ("API-016", "删除接口", "接口存在", "删除接口", "确认后删除成功", "P0"),
        ("API-017", "移动接口", "接口存在", "移动接口到其他集合", "接口移动成功", "P1"),
        ("VAR-001", "JSON提取", "接口返回JSON", "配置JSONPath提取变量", "变量提取成功", "P0"),
        ("VAR-002", "正则提取", "接口返回文本", "配置正则表达式提取", "变量提取成功", "P1"),
        ("VAR-003", "Header提取", "接口有自定义Header", "提取响应Header", "变量提取成功", "P1"),
        ("VAR-004", "变量引用", "已提取变量", "在后续请求中引用变量", "变量被正确使用", "P0"),
        ("VAR-005", "变量作用域", "多个请求", "跨请求使用变量", "变量在集合内共享", "P0"),
        ("VAR-006", "变量为空", "提取结果为空", "引用空变量", "使用空字符串", "P1"),
        ("ASSERT-001", "状态码断言", "接口已执行", "断言状态码等于200", "断言通过/失败提示", "P0"),
        ("ASSERT-002", "JSON断言", "接口返回JSON", "断言JSON中某字段值", "断言结果正确", "P0"),
        ("ASSERT-003", "响应时间断言", "接口已执行", "断言响应时间<500ms", "超过则失败", "P1"),
        ("ASSERT-004", "包含断言", "接口已执行", "断言响应包含某字符串", "包含则通过", "P0"),
        ("ASSERT-005", "正则断言", "接口已执行", "使用正则匹配", "匹配结果正确", "P1"),
        ("ASSERT-006", "多断言", "接口已执行", "添加多个断言", "全部通过才算成功", "P1"),
        ("ASSERT-007", "断言失败继续执行", "断言失败", "设置失败后继续", "记录失败但继续执行", "P2"),
        ("COL-001", "创建集合", "项目存在", "创建集合", "集合创建成功", "P0"),
        ("COL-002", "集合排序", "多个接口/集合", "拖拽排序", "顺序保存成功", "P1"),
        ("COL-003", "批量执行集合", "集合有多个接口", "点击运行集合", "按顺序执行所有接口", "P0"),
        ("COL-004", "执行中断", "集合执行中", "点击停止按钮", "执行中断，保存已执行结果", "P1"),
        ("COL-005", "集合环境选择", "多个环境", "选择执行环境", "使用选定环境执行", "P0"),
        ("COL-006", "集合变量传递", "集合有多个请求", "前置请求提取变量", "变量传递到后续请求", "P0"),
        ("COL-007", "删除集合", "集合存在", "删除集合", "确认后删除，包含接口一并删除", "P0"),
        ("HIST-001", "查看历史列表", "有执行记录", "进入历史记录页面", "显示所有执行历史", "P0"),
        ("HIST-002", "历史分页", "历史记录多", "分页查看历史", "分页正常", "P1"),
        ("HIST-003", "查看历史详情", "历史记录存在", "点击查看详情", "显示完整请求响应", "P0"),
        ("HIST-004", "查看请求体", "历史有请求体", "查看请求体详情", "正确显示请求体", "P0"),
        ("HIST-005", "查看响应体", "历史有响应", "查看响应体详情", "正确显示响应", "P0"),
        ("HIST-006", "查看断言结果", "有断言配置", "查看断言详情", "显示断言通过/失败", "P0"),
        ("HIST-007", "查看提取变量", "有变量提取", "查看提取的变量", "显示变量名和值", "P0"),
        ("HIST-008", "按接口筛选", "有多条历史", "按接口ID筛选", "只显示该接口的历史", "P1"),
        ("HIST-009", "删除单条历史", "历史记录存在", "删除单条历史", "删除成功", "P0"),
        ("HIST-010", "批量清理历史", "有历史记录", "清理30天前历史", "清理成功", "P0"),
        ("HIST-011", "关联套件执行", "有套件执行记录", "查看套件执行的历史", "显示关联的执行记录", "P1"),
    ],
    "四、通知系统": [
        ("NOTI-001", "配置飞书Webhook", "无", "配置飞书机器人Webhook", "配置保存成功", "P0"),
        ("NOTI-002", "配置企业微信", "无", "配置企业微信Webhook", "配置保存成功", "P0"),
        ("NOTI-003", "配置钉钉", "无", "配置钉钉机器人Webhook", "配置保存成功", "P0"),
        ("NOTI-004", "配置自定义Webhook", "无", "配置自定义Webhook URL", "配置保存成功", "P0"),
        ("NOTI-005", "Webhook URL无效", "无", "输入无效URL", "提示URL格式错误", "P1"),
        ("NOTI-006", "测试通知", "配置已保存", "点击测试发送", "收到测试消息", "P0"),
        ("NOTI-007", "启用/禁用通知", "配置存在", "切换启用状态", "状态切换成功", "P0"),
        ("NOTI-008", "删除配置", "配置存在", "删除通知配置", "确认后删除成功", "P0"),
        ("NOTI-009", "通知模板编辑", "配置存在", "修改通知模板", "模板保存成功", "P1"),
        ("NOTI-101", "执行完成后发送", "通知已配置", "执行API测试套件", "执行完成发送通知", "P0"),
        ("NOTI-102", "失败时发送", "通知已配置", "执行失败的测试", "失败时发送通知", "P0"),
        ("NOTI-103", "定时任务完成发送", "通知已配置", "定时任务执行完成", "发送通知", "P0"),
        ("NOTI-104", "通知发送失败", "Webhook不可达", "模拟发送失败", "记录发送失败日志", "P1"),
        ("NOTI-105", "查看通知日志", "有发送记录", "查看通知发送日志", "显示发送历史", "P0"),
    ],
    "五、定时任务": [
        ("TASK-001", "创建定时任务", "集合存在", "配置定时任务，选择执行集合", "任务创建成功", "P0"),
        ("TASK-002", "Cron表达式", "无", "配置Cron表达式", "显示下次执行时间", "P0"),
        ("TASK-003", "修改定时任务", "任务存在", "修改执行时间", "修改成功", "P0"),
        ("TASK-004", "立即执行", "任务存在", "点击立即执行", "立即执行一次", "P0"),
        ("TASK-005", "暂停任务", "任务运行中", "点击暂停", "任务暂停，不再自动执行", "P0"),
        ("TASK-006", "恢复任务", "任务已暂停", "点击恢复", "任务恢复自动执行", "P0"),
        ("TASK-007", "删除任务", "任务存在", "删除定时任务", "确认后删除成功", "P0"),
        ("TASK-008", "执行记录", "任务有执行", "查看执行记录", "显示执行历史", "P0"),
        ("TASK-101", "Cron格式错误", "无", "输入错误Cron表达式", "提示Cron格式错误", "P1"),
        ("TASK-102", "过去时间", "无", "设置过去时间为执行时间", "提示时间已过期", "P1"),
        ("TASK-103", "并发执行", "任务执行时间长", "设置短周期任务", "避免重复执行", "P1"),
    ],
    "六、操作日志": [
        ("LOG-001", "查看日志列表", "有操作记录", "进入操作日志页面", "显示操作日志列表", "P0"),
        ("LOG-002", "按类型筛选", "有多种操作类型", "筛选创建操作", "只显示创建操作", "P1"),
        ("LOG-003", "按资源类型筛选", "有多种资源", "筛选项目操作", "只显示项目相关操作", "P1"),
        ("LOG-004", "按用户筛选", "有多个用户操作", "按用户筛选", "显示该用户的所有操作", "P1"),
        ("LOG-005", "查看日志详情", "日志存在", "点击查看详情", "显示完整操作信息", "P0"),
        ("LOG-006", "日志分页", "日志记录多", "分页查看", "分页正常", "P1"),
        ("LOG-101", "删除单条日志", "日志存在", "删除单条日志", "删除成功", "P0"),
        ("LOG-102", "清理旧日志", "有30天前日志", "清理30天前日志", "清理成功", "P0"),
        ("LOG-103", "清理未来日期", "无", "尝试清理未来日志", "提示无日志可清理", "P2"),
    ],
    "七、边界测试": [
        ("BOUND-001", "超长输入", "无", "输入超长字符串(>10000字符)", "正确处理或截断", "P1"),
        ("BOUND-002", "特殊字符", "无", "输入特殊字符`<>&"'`", "正确转义存储", "P0"),
        ("BOUND-003", "中文输入", "无", "输入中文内容", "正确保存和显示", "P0"),
        ("BOUND-004", "空格处理", "无", "输入首尾空格", "自动去除或保留一致", "P1"),
        ("BOUND-005", "空字符串", "无", "提交空字符串", "提示必填或允许", "P1"),
        ("BOUND-006", "负数输入", "数字字段", "输入负数", "正确处理", "P1"),
        ("BOUND-007", "零值输入", "数字字段", "输入0", "正确处理", "P1"),
        ("BOUND-008", "超大数字", "数字字段", "输入超大数字", "正确处理或溢出提示", "P2"),
        ("CONC-001", "同时登录", "无", "同一账号多处同时登录", "都能成功或后者挤掉前者", "P1"),
        ("CONC-002", "同时编辑", "两人同时编辑", "两人同时修改同一接口", "后提交者覆盖或提示冲突", "P1"),
        ("CONC-003", "快速提交", "无", "快速连续点击提交按钮", "只提交一次或队列处理", "P1"),
        ("CONC-004", "大量并发请求", "无", "同时发起大量请求", "系统稳定不过载", "P2"),
        ("NET-001", "网络中断", "请求执行中", "断开网络", "显示网络错误", "P0"),
        ("NET-002", "网络恢复", "网络中断后恢复", "恢复网络", "继续正常工作", "P1"),
        ("NET-003", "弱网环境", "网络慢", "在弱网环境下操作", "正确处理超时", "P2"),
        ("NET-004", "DNS解析失败", "目标域名无效", "请求无效域名", "显示DNS错误", "P1"),
    ],
    "八、安全测试": [
        ("SEC-001", "无Token访问", "未登录", "直接访问API", "返回401未授权", "P0"),
        ("SEC-002", "伪造Token", "无", "使用伪造Token访问", "返回401未授权", "P0"),
        ("SEC-003", "过期Token", "Token已过期", "使用过期Token", "返回401，提示过期", "P0"),
        ("SEC-004", "越权访问", "普通用户", "访问管理员接口", "返回403无权限", "P0"),
        ("SEC-005", "跨项目访问", "无权限项目", "访问无权限项目数据", "返回404或403", "P0"),
        ("SEC-101", "SQL注入", "输入字段", "输入SQL语句", "被转义，不执行SQL", "P0"),
        ("SEC-102", "XSS攻击", "输入字段", "输入`<script>`标签", "被转义，不执行脚本", "P0"),
        ("SEC-103", "密码明文传输", "登录时", "抓包查看密码", "密码加密传输", "P0"),
        ("SEC-104", "Token泄露", "无", "Token在URL中", "Token不在URL，在Header", "P1"),
    ],
    "九、兼容性测试": [
        ("COMP-001", "Chrome浏览器", "Chrome最新", "所有功能测试", "正常工作", "P0"),
        ("COMP-002", "Firefox浏览器", "Firefox最新", "所有功能测试", "正常工作", "P1"),
        ("COMP-003", "Edge浏览器", "Edge最新", "所有功能测试", "正常工作", "P1"),
        ("COMP-004", "Safari浏览器", "Safari最新", "所有功能测试", "正常工作", "P2"),
        ("COMP-005", "手机端", "iOS/Android", "访问系统", "基本功能可用", "P2"),
        ("COMP-006", "1024x768分辨率", "低分辨率", "查看界面", "界面正常显示", "P1"),
        ("COMP-007", "4K分辨率", "高分辨率", "查看界面", "界面正常显示", "P2"),
    ],
    "十、性能测试": [
        ("PERF-001", "登录响应时间", "单用户登录", "测量登录响应时间", "<2秒", "P1"),
        ("PERF-002", "列表加载时间", "100条数据", "加载列表", "<3秒", "P1"),
        ("PERF-003", "接口执行时间", "简单请求", "执行一个简单GET请求", "<5秒", "P0"),
        ("PERF-004", "并发用户", "10用户同时", "模拟10用户同时操作", "系统正常响应", "P2"),
        ("PERF-005", "内存占用", "长时间使用", "使用1小时", "内存不持续增长", "P2"),
    ],
}

def create_sheet(wb, sheet_name, cases):
    """创建单个模块的工作表"""
    ws = wb.create_sheet(title=sheet_name)

    # 设置列宽
    column_widths = {
        'A': 12,   # 用例编号
        'B': 20,   # 用例名称
        'C': 15,   # 前置条件
        'D': 35,   # 测试步骤
        'E': 35,   # 预期结果
        'F': 8,    # 优先级
        'G': 12,   # 测试日期
        'H': 10,   # 测试人员
        'I': 12,   # 测试结果
        'J': 30,   # 备注
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 设置行高
    ws.row_dimensions[1].height = 30

    # 写入表头
    headers = ["用例编号", "用例名称", "前置条件", "测试步骤", "预期结果", "优先级", "测试日期", "测试人员", "测试结果", "备注"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入测试用例数据
    for row_idx, case in enumerate(cases, 2):
        for col_idx, value in enumerate(case, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = wrap_alignment

        # 设置优先级颜色
        priority = case[5]
        priority_cell = ws.cell(row=row_idx, column=6)
        if priority == "P0":
            priority_cell.fill = p0_fill
        elif priority == "P1":
            priority_cell.fill = p1_fill
        elif priority == "P2":
            priority_cell.fill = p2_fill
        priority_cell.alignment = center_alignment

        # 设置测试结果列的下拉选项和默认样式
        result_cell = ws.cell(row=row_idx, column=9, value="")
        result_cell.border = thin_border

    return ws

# 删除默认sheet
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

# 创建各模块工作表
for module_name, cases in test_cases.items():
    # 简化工作表名称（去掉序号）
    short_name = module_name.split("、")[1] if "、" in module_name else module_name
    create_sheet(wb, short_name, cases)

# 创建汇总页
ws_summary = wb.create_sheet(title="测试汇总", index=0)

# 设置列宽
summary_widths = {
    'A': 25,
    'B': 15,
    'C': 15,
    'D': 15,
    'E': 15,
    'F': 15,
    'G': 20,
}
for col, width in summary_widths.items():
    ws_summary.column_dimensions[col].width = width

# 写入标题
title_cell = ws_summary.cell(row=1, column=1, value="TestHub Java版本 系统测试用例汇总")
title_cell.font = Font(bold=True, size=14)
ws_summary.merge_cells('A1:G1')
title_cell.alignment = center_alignment

# 写入汇总表头
summary_headers = ["模块名称", "用例总数", "P0用例", "P1用例", "P2用例", "已完成", "完成率"]
for col, header in enumerate(summary_headers, 1):
    cell = ws_summary.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# 写入汇总数据
row = 4
total_cases = 0
total_p0 = 0
total_p1 = 0
total_p2 = 0

for module_name, cases in test_cases.items():
    short_name = module_name.split("、")[1] if "、" in module_name else module_name
    p0_count = sum(1 for c in cases if c[5] == "P0")
    p1_count = sum(1 for c in cases if c[5] == "P1")
    p2_count = sum(1 for c in cases if c[5] == "P2")
    total = len(cases)

    total_cases += total
    total_p0 += p0_count
    total_p1 += p1_count
    total_p2 += p2_count

    data = [module_name, total, p0_count, p1_count, p2_count, 0, "0%"]
    for col, value in enumerate(data, 1):
        cell = ws_summary.cell(row=row, column=col, value=value)
        cell.border = thin_border
        cell.alignment = center_alignment if col > 1 else Alignment(horizontal="left", vertical="center")

    row += 1

# 写入总计行
total_data = ["总计", total_cases, total_p0, total_p1, total_p2, 0, "0%"]
for col, value in enumerate(total_data, 1):
    cell = ws_summary.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    cell.border = thin_border
    cell.alignment = center_alignment if col > 1 else Alignment(horizontal="left", vertical="center")

# 创建测试执行记录页
ws_record = wb.create_sheet(title="执行记录", index=1)

record_widths = {'A': 15, 'B': 15, 'C': 15, 'D': 10, 'E': 10, 'F': 10, 'G': 30}
for col, width in record_widths.items():
    ws_record.column_dimensions[col].width = width

record_title = ws_record.cell(row=1, column=1, value="测试执行记录")
record_title.font = Font(bold=True, size=14)
ws_record.merge_cells('A1:G1')
record_title.alignment = center_alignment

record_headers = ["日期", "测试人员", "测试用例数", "通过", "失败", "阻塞", "备注"]
for col, header in enumerate(record_headers, 1):
    cell = ws_record.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# 预设5行空白记录
for r in range(4, 9):
    for c in range(1, 8):
        cell = ws_record.cell(row=r, column=c, value="")
        cell.border = thin_border

# 创建缺陷记录页
ws_bug = wb.create_sheet(title="缺陷记录", index=2)

bug_widths = {'A': 12, 'B': 15, 'C': 40, 'D': 12, 'E': 12, 'F': 30}
for col, width in bug_widths.items():
    ws_bug.column_dimensions[col].width = width

bug_title = ws_bug.cell(row=1, column=1, value="缺陷记录")
bug_title.font = Font(bold=True, size=14)
ws_bug.merge_cells('A1:F1')
bug_title.alignment = center_alignment

bug_headers = ["缺陷ID", "用例编号", "缺陷描述", "严重程度", "状态", "备注"]
for col, header in enumerate(bug_headers, 1):
    cell = ws_bug.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# 预设10行空白记录
for r in range(4, 14):
    for c in range(1, 7):
        cell = ws_bug.cell(row=r, column=c, value="")
        cell.border = thin_border

# 添加使用说明
ws_guide = wb.create_sheet(title="使用说明", index=3)

guide_content = [
    ("TestHub Java版本 系统测试用例", Font(bold=True, size=16)),
    ("", None),
    ("一、工作表说明", Font(bold=True, size=12)),
    ("测试汇总: 显示各模块测试用例统计", None),
    ("执行记录: 记录每次测试执行的情况", None),
    ("缺陷记录: 记录测试过程中发现的缺陷", None),
    ("其他工作表: 按模块分类的详细测试用例", None),
    ("", None),
    ("二、测试结果填写说明", Font(bold=True, size=12)),
    ("测试结果列(P列): 填写以下值之一:", None),
    ("  - 通过: 测试执行结果符合预期", None),
    ("  - 失败: 测试执行结果与预期不符", None),
    ("  - 阻塞: 因环境或前置条件问题无法执行", None),
    ("  - N/A: 不适用该用例", None),
    ("", None),
    ("三、优先级说明", Font(bold=True, size=12)),
    ("  - P0: 核心功能，必须通过 (黄色)", None),
    ("  - P1: 重要功能，建议通过 (绿色)", None),
    ("  - P2: 一般功能，可选通过 (蓝色)", None),
    ("", None),
    ("四、测试流程建议", Font(bold=True, size=12)),
    ("1. 先完成P0级别用例测试", None),
    ("2. 再完成P1级别用例测试", None),
    ("3. 最后根据时间完成P2级别用例测试", None),
    ("4. 每日记录测试执行情况", None),
    ("5. 发现缺陷及时记录到缺陷记录表", None),
]

for row, (content, font) in enumerate(guide_content, 1):
    cell = ws_guide.cell(row=row, column=1, value=content)
    if font:
        cell.font = font
    ws_guide.column_dimensions['A'].width = 50

# 保存文件
output_path = "D:/Project/testhub_platform/testhub-java/docs/TestHub-Java系统测试用例.xlsx"
wb.save(output_path)
print(f"Excel文件已生成: {output_path}")
print(f"总计 {total_cases} 个测试用例")
print(f"P0用例: {total_p0} 个")
print(f"P1用例: {total_p1} 个")
print(f"P2用例: {total_p2} 个")
